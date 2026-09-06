from openpyxl import load_workbook

from helpers.excel_export import export_rows_to_excel
from workers.db_worker import DbWorker


def test_export_rows_to_excel_preserves_unicode_and_adds_extension(tmp_path):
    output = export_rows_to_excel(
        ["Название", "SKU"],
        [["Цепь", "ABC-1"]],
        tmp_path / "export",
    )

    assert output == tmp_path / "export.xlsx"
    workbook = load_workbook(output)
    worksheet = workbook["Товары"]
    assert [cell.value for cell in worksheet[1]] == ["Название", "SKU"]
    assert worksheet["A2"].value == "Цепь"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:B2"


def test_export_rows_to_excel_keeps_fractional_values_numeric(tmp_path):
    output = export_rows_to_excel(
        ["Цена", "Вес (кг)"],
        [[496.94, 0.01]],
        tmp_path / "export.xlsx",
    )

    worksheet = load_workbook(output)["Товары"]

    assert worksheet["A2"].value == 496.94
    assert worksheet["A2"].number_format == "#,##0.00"
    assert worksheet["B2"].value == 0.01
    assert worksheet["B2"].number_format == "#,##0.00"


def test_db_worker_emits_finished_result():
    worker = DbWorker(lambda value: value + 1, 1)
    results = []
    worker.signals.finished.connect(results.append)

    worker.run()

    assert results == [2]


def test_db_worker_emits_error():
    worker = DbWorker(lambda: (_ for _ in ()).throw(ValueError("broken")))
    errors = []
    worker.signals.error.connect(errors.append)

    worker.run()

    assert len(errors) == 1
    assert str(errors[0]) == "broken"
