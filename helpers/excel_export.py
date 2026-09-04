from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app_logging import logger


def export_rows_to_excel(
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    file_path: str | Path,
) -> Path:
    """Export tabular data to an Excel workbook with Unicode support."""
    output_path = Path(file_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Товары"

    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(list(row))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(width, 60)

    try:
        workbook.save(output_path)
    except Exception:
        logger.exception("Ошибка экспорта Excel: %s", output_path)
        raise
    return output_path
